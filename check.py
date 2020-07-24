from bs4 import BeautifulSoup
from openpyxl import load_workbook
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import time
from datetime import datetime
from twilio.rest import Client
import requests

#load excel sheet to export price data to
workbook = load_workbook("checkPrice.xlsx")
#establish twilio client for sms capabilities
client = Client("xxxxxxxxxxxxx", "xxxxxxxxxxxxxxxxxx")

#create loop variable to allow for an infinite loop
loop = 1
while loop == 1: #will always evaluate to true
    now = datetime.now()
    date_time = now.strftime("%m-%d-%Y")
    sheet = workbook.create_sheet(date_time) #create sheet with date and time -- this allows for me to track the day-by-day data to compare them
    #user agent header needed to access site
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.89 Safari/537.36'}
    #create http client
    page = requests.get("https://www.urbanoutfitters.com/jeans-for-women?size_class_group=Waist~30", headers=headers)

    #initialize lists for the sheet columns
    products = [];
    prices = [];
    oldPrices = [];

    #parse through the page
    soup = BeautifulSoup(page.content, features="html.parser")

    #find all instances of the class where the price and name information is
    for a in soup.findAll('div', {'class':'c-pwa-product-tile'}):
        #store product name
        name=a.find('p', {'class':'c-pwa-product-tile__heading'})
        #store product price
        price=a.find('span', {'class':'c-pwa-product-price__current'})
        #store original product price (if the item is on sale this will be a value, if not it will be NONE
        oldPrice = a.find('span', {'class':'c-pwa-product-price__original'})
        #add new name and price to lists
        products.append(name.text)
        prices.append(price.text)
        #if the item is not on sale, the sale column will be blank
        if (oldPrice == None):
            oldPrices.append(" ")
        else:
            #if an item is on sale, add the old price to the list and send myself a text
            client.api.account.messages.create(to="+11111111111",from_="+1111111111111",body="An item is on sale!")
            oldPrices.append("SALE: " + oldPrice.text)

    #create a data frame with the price and name information
    df = pd.DataFrame({'Product Name':products,'Price':prices,'Sale -- Original Price':oldPrices})
    #add each data row to the excel sheet
    for row in dataframe_to_rows(df, index=False, header=True):
        sheet.append(row)

    #save the excel document
    workbook.save("C:/Users/savan/Fun Coding Stuff/CheckPrice/checkPrice.xlsx")
    #wait one day (86400 seconds) before running again, allows me to check the prices once per day
    time.sleep(86400)

